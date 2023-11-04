from openpyxl import load_workbook
import pandas as pd
import os
from tqdm import tqdm
from main import LOAD_PATH
from regions import regions_in_federal

def convert_excel():
    os.chdir(LOAD_PATH)

    # отбираем xlsx файлы в директории и добавляем в список
    list_of_files = [file for file in os.listdir() if file.endswith('.xlsx')]

    # создаем два списка для хранения датасетов
    list_of_df_long = []
    list_of_df_short = []

    # обрабатываем файлы
    print('Processing files:')
    for file in tqdm(list_of_files, colour='#FC46AA'):
        ind = load_workbook(file).active['A2'].value.lower().split(' ')[0] # читаем показатель
        year = load_workbook(file).active['B3'].value # читаем год
        loc = load_workbook(file).active['B4'].value # читаем локализацию
        df = pd.read_excel(file, header=6, decimal=',') # столбцы в 7 строке, разделитель - запятая
        df['ind'] = ind # добавляем колонку показатель
        df['year'] = year # добавляем колонку год
        df['loc'] = loc # добавляем колонку локализация
        if len(df.columns) == 8: # если столбцов 8
            old_cols = df.columns.tolist()
            new_cols = ['region', 'total_abs', 'total_rude', 'total_standart', 'total_mistake']
            cols_dict = {k: v for k, v in zip(old_cols, new_cols)}
            df = df.rename(columns=cols_dict) # меняем название колонок
            list_of_df_short.append(df)
        else:
            list_of_df_long.append(df)

    # объединяем датасеты
    df_long = pd.concat(list_of_df_long)
    df_short = pd.concat(list_of_df_short)

    # меняем названия колонок в широкой версии датасета
    new_long_columns = ['region', 'all_abs', 'all_rude', 'all_standart', 'all_mistake', 
                        'men_abs', 'men_rude', 'men_standart', 'men_mistake',
                    'women_abs', 'women_rude', 'women_standart', 'women_mistake', 'ind', 'year', 'loc']
    old_long_columns = df_long.columns.tolist()
    long_col_dict = {k: v for k, v in zip(old_long_columns, new_long_columns)}
    df_long = df_long.rename(columns=long_col_dict)

    # чистим пробелы
    df_short.region = df_short.region.str.strip()
    df_long.region = df_long.region.str.strip()

    # список фраз для удаления из датасетов
    forbidden_values = ['РОССИЯ', 'ЦЕНТРАЛЬНЫЙ ФО', 'СЕВЕРО-ЗАПАДНЫЙ ФО', 'ЮЖНЫЙ ФО', 
                        'СЕВЕРО-КАВКАЗСКИЙ ФО', 'ПРИВОЛЖСКИЙ ФО', 'УРАЛЬСКИЙ ФО', 
                        'СИБИРСКИЙ ФО', 'ДАЛЬНЕВОСТОЧНЫЙ ФО']

    # чистим датасеты от агрегированных значений по регионам и РФ
    df_long = df_long[~df_long.region.isin(forbidden_values)]
    df_short = df_short[~df_short.region.isin(forbidden_values)]

    
    def make_fed(cell):
        'Функция для добавления столбца Федеральный округ по округам'
        for k, v in regions_in_federal.items():
            if cell in v:
                return k

    # добавляем федеральный округ в датасеты
    df_long['federal'] = df_long.region.apply(make_fed)
    df_short['federal'] = df_short.region.apply(make_fed)

    # заполняем пропущенное значение года
    df_long.year = df_long.year.fillna(2021)

    # меняем тип данных столбца год на int
    df_long.year = df_long.year.astype('int')
    df_short.year = df_short.year.astype('int')
    print(f'{df_long.shape}')
    print(f'{df_short.shape}')
    # сохраняем датасеты
    df_long.to_csv('df_long.csv', encoding='utf8')
    df_short.to_csv('df_short.csv', encoding='utf8')
    print('Files created. Script over.')